"""
General utility functions.
"""
import sys
import os
import csv
import re
from time import time
from datetime import datetime, timedelta
from functools import wraps, lru_cache

from vnpy_pro.trader.object import BarData
from vnpy.trader.constant import Exchange, Interval


def func_time(over_ms: int = 0):
    """
    简单记录执行时间
    :param :over_ms 超过多少毫秒, 提示信息
    :return:
    """

    def run(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            start = time()
            result = func(*args, **kwargs)
            end = time()
            execute_ms = (int(round(end * 1000))) - (int(round(start * 1000)))
            if execute_ms > over_ms:
                print('{} took {} ms'.format(func.__qualname__, execute_ms))
            return result

        return wrapper

    return run


@lru_cache()
def get_underlying_symbol(symbol: str):
    """
    取得合约的短号.  rb2005 => rb
    :param symbol:
    :return: 短号
    """
    # 套利合约
    if symbol.find(' ') != -1:
        # 排除SP SPC SPD
        s = symbol.split(' ')
        if len(s) < 2:
            return symbol
        symbol = s[1]

        # 只提取leg1合约
        if symbol.find('&') != -1:
            s = symbol.split('&')
            if len(s) < 2:
                return symbol
            symbol = s[0]

    p = re.compile(r"([A-Z]+)[0-9]+", re.I)
    underlying_symbol = p.match(symbol)

    if underlying_symbol is None:
        return symbol

    return underlying_symbol.group(1)


@lru_cache()
def get_stock_exchange(code, vn=True):
    """根据股票代码，获取交易所"""
    # vn：取EXCHANGE_SSE 和 EXCHANGE_SZSE
    code = str(code)
    if len(code) < 6:
        return ''

    market_id = 0  # 缺省深圳
    code = str(code)
    if code[0] in ['5', '6', '9'] or code[:3] in ["009", "126", "110", "201", "202", "203", "204"]:
        market_id = 1  # 上海
    try:
        from vnpy.trader.constant import Exchange
        if vn:
            return Exchange.SSE.value if market_id == 1 else Exchange.SZSE.value
        else:
            return 'XSHG' if market_id == 1 else 'XSHE'
    except Exception as ex:
        print(u'加载数据异常:{}'.format(str(ex)))

    return ''


@lru_cache()
def get_full_symbol(symbol: str):
    """
    获取全路径得合约名称, MA005 => MA2005, j2005 => j2005
    """
    if symbol.endswith('SPD'):
        return symbol

    underlying_symbol = get_underlying_symbol(symbol)
    if underlying_symbol == symbol:
        return symbol

    symbol_month = symbol.replace(underlying_symbol, '')
    if len(symbol_month) == 3:
        if symbol_month[0] == '0':
            # 支持2020年合约
            return '{0}2{1}'.format(underlying_symbol, symbol_month)
        else:
            return '{0}1{1}'.format(underlying_symbol, symbol_month)
    else:
        return symbol


def get_real_symbol_by_exchange(full_symbol, vn_exchange):
    """根据交易所，返回真实合约"""
    if vn_exchange == Exchange.CFFEX:
        return full_symbol.upper()

    if vn_exchange in [Exchange.DCE, Exchange.SHFE, Exchange.INE]:
        return full_symbol.lower()

    if vn_exchange == Exchange.CZCE:
        underlying_symbol = get_underlying_symbol(full_symbol).upper()
        yearmonth_len = len(full_symbol) - len(underlying_symbol) - 1
        return underlying_symbol.upper() + full_symbol[-yearmonth_len:]

    return full_symbol


def get_trading_date(dt: datetime = None):
    """
    根据输入的时间，返回交易日的日期
    :param dt:
    :return:
    """
    if dt is None:
        dt = datetime.now()

    if dt.isoweekday() in [6, 7]:
        # 星期六,星期天=>星期一
        return (dt + timedelta(days=8 - dt.isoweekday())).strftime('%Y-%m-%d')

    if dt.hour >= 20:
        if dt.isoweekday() == 5:
            # 星期五=》星期一
            return (dt + timedelta(days=3)).strftime('%Y-%m-%d')
        else:
            # 第二天
            return (dt + timedelta(days=1)).strftime('%Y-%m-%d')
    else:
        return dt.strftime('%Y-%m-%d')


def format_number(n):
    """格式化数字到字符串"""
    rn = round(n, 2)  # 保留两位小数
    return format(rn, ',')  # 加上千分符


def print_dict(d: dict):
    """返回dict的字符串类型"""
    return '\n'.join([f'{key}:{d[key]}' for key in sorted(d.keys())])


def append_data(file_name: str, dict_data: dict, field_names=None):
    """
    添加数据到csv文件中
    :param field_names:
    :param file_name:  csv的文件全路径
    :param dict_data:  OrderedDict
    :return:
    """
    if field_names is None:
        field_names = []
    dict_fieldnames = sorted(list(dict_data.keys())) if len(field_names) == 0 else field_names

    try:
        if not os.path.exists(file_name):
            print(u'create csv file:{}'.format(file_name))
            with open(file_name, 'a', encoding='utf8', newline='\n') as csvWriteFile:
                writer = csv.DictWriter(f=csvWriteFile, fieldnames=dict_fieldnames, dialect='excel')
                print(u'write csv header:{}'.format(dict_fieldnames))
                writer.writeheader()
                writer.writerow(dict_data)
        else:
            with open(file_name, 'a', encoding='utf8', newline='\n') as csvWriteFile:
                writer = csv.DictWriter(f=csvWriteFile, fieldnames=dict_fieldnames, dialect='excel',
                                        extrasaction='ignore')
                writer.writerow(dict_data)
    except Exception as ex:
        print(u'append_data exception:{}'.format(str(ex)), file=sys.stderr)


def import_module_by_str(import_module_name):
    """
    动态导入模块/函数
    :param import_module_name:
    :return:
    """
    import traceback
    from importlib import import_module, reload

    # 参数检查
    if len(import_module_name) == 0:
        print('import_module_by_str parameter error,return None')
        return None

    print('trying to import {}'.format(import_module_name))
    try:
        import_name = str(import_module_name).replace(':', '.')
        modules = import_name.split('.')
        if len(modules) == 1:
            mod = import_module(modules[0])
            return mod
        else:
            loaded_modules = '.'.join(modules[0:-1])
            print('import {}'.format(loaded_modules))
            mod = import_module(loaded_modules)

            comp = modules[-1]
            if not hasattr(mod, comp):
                loaded_modules = '.'.join([loaded_modules, comp])
                print('realod {}'.format(loaded_modules))
                mod = reload(loaded_modules)
            else:
                print('from {} import {}'.format(loaded_modules, comp))
                mod = getattr(mod, comp)
            return mod

    except Exception as ex:
        print('import {} fail,{},{}'.format(import_module_name, str(ex), traceback.format_exc()))

        return None


def save_df_to_excel(file_name, sheet_name, df):
    """
    保存dataframe到execl
    :param file_name: 保存的excel文件名
    :param sheet_name: 保存的sheet
    :param df: dataframe
    :return: True/False
    """
    if file_name is None or sheet_name is None or df is None:
        return False

    # ----------------------------- 扩展的功能 ---------
    try:
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        # from openpyxl.drawing.image import Image
    except:  # noqa
        print(u'can not import openpyxl', file=sys.stderr)

    if 'openpyxl' not in sys.modules:
        print(u'can not import openpyxl', file=sys.stderr)
        return False

    try:
        ws = None

        try:
            # 读取文件
            wb = openpyxl.load_workbook(file_name)
        except:  # noqa
            # 创建一个excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
        try:
            # 定位WorkSheet
            if ws is None:
                ws = wb[sheet_name]
        except:  # noqa
            # 创建一个WorkSheet
            ws = wb.create_sheet()
            ws.title = sheet_name

        rows = dataframe_to_rows(df)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Save the workbook
        wb.save(file_name)
        wb.close()
    except Exception as ex:
        import traceback
        print(u'save_df_to_excel exception:{}'.format(str(ex)), traceback.format_exc(), file=sys.stderr)


def save_text_to_excel(file_name, sheet_name, text):
    """
    保存文本文件到excel
    :param file_name:
    :param sheet_name:
    :param text:
    :return:
    """
    if file_name is None or len(sheet_name) == 0 or len(text) == 0:
        return False

    # ----------------------------- 扩展的功能 ---------
    try:
        import openpyxl
        # from openpyxl.utils.dataframe import dataframe_to_rows
        # from openpyxl.drawing.image import Image
    except:  # noqa
        print(u'can not import openpyxl', file=sys.stderr)

    if 'openpyxl' not in sys.modules:
        return False

    try:
        ws = None
        try:
            # 读取文件
            wb = openpyxl.load_workbook(file_name)
        except:  # noqa
            # 创建一个excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
        try:
            # 定位WorkSheet
            if ws is None:
                ws = wb[sheet_name]
        except:  # noqa
            # 创建一个WorkSheet
            ws = wb.create_sheet()
            ws.title = sheet_name

        # 设置宽度，自动换行方式
        ws.column_dimensions["A"].width = 120
        ws['A2'].alignment = openpyxl.styles.Alignment(wrapText=True)
        ws['A2'].value = text

        # Save the workbook
        wb.save(file_name)
        wb.close()
        return True
    except Exception as ex:
        import traceback
        print(u'save_text_to_excel exception:{}'.format(str(ex)), traceback.format_exc(), file=sys.stderr)
        return False


def save_images_to_excel(file_name, sheet_name, image_names):
    """
    # 保存图形文件到excel
    :param file_name: excel文件名
    :param sheet_name: workSheet
    :param image_names: 图像文件名列表
    :return:
    """
    if file_name is None or len(sheet_name) == 0 or len(image_names) == 0:
        return False
    # ----------------------------- 扩展的功能 ---------
    try:
        import openpyxl
        # from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.drawing.image import Image
    except Exception as ex:
        print(f'can not import openpyxl:{str(ex)}', file=sys.stderr)

    if 'openpyxl' not in sys.modules:
        return False
    try:
        ws = None

        try:
            # 读取文件
            wb = openpyxl.load_workbook(file_name)
        except:  # noqa
            # 创建一个excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
        try:
            # 定位WorkSheet
            if ws is None:
                ws = wb[sheet_name]
        except Exception as ex:  # noqa
            # 创建一个WorkSheet
            ws = wb.create_sheet()
            ws.title = sheet_name

        i = 1

        for image_name in image_names:
            try:
                # 加载图形文件
                img1 = Image(image_name)

                cell_id = 'A{0}'.format(i)
                ws[cell_id].value = image_name
                cell_id = 'A{0}'.format(i + 1)

                i += 30

                # 添加至对应的WorkSheet中
                ws.add_image(img1, cell_id)
            except Exception as ex:
                print('exception loading image {}, {}'.format(image_name, str(ex)), file=sys.stderr)
                return False

        # Save the workbook
        wb.save(file_name)
        wb.close()
        return True
    except Exception as ex:
        import traceback
        print(u'save_images_to_excel exception:{}'.format(str(ex)), traceback.format_exc(), file=sys.stderr)
        return False


def display_dual_axis(df, columns1, columns2=[], invert_yaxis1=False, invert_yaxis2=False, file_name=None,
                      sheet_name=None,
                      image_name=None):
    """
    显示(保存)双Y轴的走势图
    :param df: DataFrame
    :param columns1:  y1轴
    :param columns2: Y2轴
    :param invert_yaxis1: Y1 轴反转
    :param invert_yaxis2: Y2 轴翻转
    :param file_name:   保存的excel 文件名称
    :param sheet_name:  excel 的sheet
    :param image_name:  保存的image 文件名
    :return:
    """

    import matplotlib
    import matplotlib.pyplot as plt
    matplotlib.rcParams['figure.figsize'] = (20.0, 10.0)

    df1 = df[columns1]
    df1.index = list(range(len(df)))
    fig, ax1 = plt.subplots()
    if invert_yaxis1:
        ax1.invert_yaxis()
    ax1.plot(df1)

    if len(columns2) > 0:
        df2 = df[columns2]
        df2.index = list(range(len(df)))
        ax2 = ax1.twinx()
        if invert_yaxis2:
            ax2.invert_yaxis()
        ax2.plot(df2)

    # 修改x轴得label为时间
    xt = ax1.get_xticks()
    xt2 = [df.index[int(i)] for i in xt[1:-2]]
    xt2.insert(0, '')
    xt2.append('')
    ax1.set_xticklabels(xt2)

    # 是否保存图片到文件
    if image_name is not None:
        fig = plt.gcf()
        fig.savefig(image_name, bbox_inches='tight')

        # 插入图片到指定的excel文件sheet中并保存excel
        if file_name is not None and sheet_name is not None:
            save_images_to_excel(file_name, sheet_name, [image_name])
    else:
        plt.show()


def get_bars(csv_file: str,
             symbol: str,
             exchange: Exchange,
             start_date: datetime = None,
             end_date: datetime = None, ):
    """
    获取bar
    数据存储目录: 项目/bar_data
    :param csv_file: csv文件路径
    :param symbol: 合约
    :param exchange 交易所
    :param start_date: datetime
    :param end_date: datetime
    :return:
    """
    bars = []

    import csv
    with open(file=csv_file, mode='r', encoding='utf8', newline='\n') as f:
        reader = csv.DictReader(f)

        count = 0

        for item in reader:

            dt = datetime.strptime(item['datetime'], '%Y-%m-%d %H:%M:%S')
            if start_date:
                if dt < start_date:
                    continue
            if end_date:
                if dt > end_date:
                    break

            bar = BarData(
                symbol=symbol,
                exchange=exchange,
                datetime=dt,
                interval=Interval.MINUTE,
                volume=float(item['volume']),
                open_price=float(item['open']),
                high_price=float(item['high']),
                low_price=float(item['low']),
                close_price=float(item['close']),
                open_interest=float(item['open_interest']),
                trading_day=item['trading_day'],
                gateway_name="Tdx",
            )

            bars.append(bar)

            # do some statistics
            count += 1

    return bars
